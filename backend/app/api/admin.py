from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from sqlalchemy.orm import selectinload
from datetime import datetime, date, timedelta
from typing import Optional
import io

from app.database import get_db
from app.config import settings
from app.models.user import User
from app.models.student import Student
from app.models.teacher import Teacher
from app.models.group import Group
from app.models.direction import Direction
from app.models.subject import Subject
from app.models.schedule import Schedule
from app.models.lesson import Lesson
from app.models.attendance import Attendance
from app.api.auth import get_current_user

router = APIRouter(prefix="/admin", tags=["admin"])


async def check_admin(current_user: User, db: AsyncSession):
    """Admin tekshirish"""
    admin_ids = [int(x.strip()) for x in settings.ADMIN_IDS.split(',') if x.strip()]
    if current_user.telegram_id not in admin_ids:
        raise HTTPException(status_code=403, detail="Admin huquqi kerak")
    return True


@router.get("/stats")
async def get_stats(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Umumiy statistika"""
    await check_admin(current_user, db)
    
    students_count = await db.scalar(select(func.count(Student.id)))
    teachers_count = await db.scalar(select(func.count(Teacher.id)))
    groups_count = await db.scalar(select(func.count(Group.id)))
    lessons_count = await db.scalar(select(func.count(Lesson.id)))
    attendance_count = await db.scalar(select(func.count(Attendance.id)))
    
    # Bugungi statistika
    today = date.today()
    today_lessons = await db.scalar(
        select(func.count(Lesson.id)).where(Lesson.date == today)
    )
    today_attendance = await db.scalar(
        select(func.count(Attendance.id))
        .join(Lesson)
        .where(Lesson.date == today)
    )
    
    return {
        "total_students": students_count or 0,
        "total_teachers": teachers_count or 0,
        "total_groups": groups_count or 0,
        "total_lessons": lessons_count or 0,
        "total_attendance": attendance_count or 0,
        "today_lessons": today_lessons or 0,
        "today_attendance": today_attendance or 0
    }


@router.get("/students")
async def get_all_students(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Barcha talabalar"""
    await check_admin(current_user, db)
    
    result = await db.execute(
        select(Student)
        .options(
            selectinload(Student.user),
            selectinload(Student.group).selectinload(Group.direction)
        )
        .order_by(Student.id)
    )
    students = result.scalars().all()
    
    return [{
        "id": s.id,
        "user_id": s.user_id,
        "full_name": s.user.full_name if s.user else None,
        "student_id": s.student_id,
        "group_name": s.group.name if s.group else None,
        "direction_name": s.group.direction.name if s.group and s.group.direction else None,
        "created_at": s.created_at.isoformat() if s.created_at else None
    } for s in students]


@router.get("/teachers")
async def get_all_teachers(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Barcha o'qituvchilar"""
    await check_admin(current_user, db)
    
    result = await db.execute(
        select(Teacher)
        .options(selectinload(Teacher.user))
        .order_by(Teacher.id)
    )
    teachers = result.scalars().all()
    
    return [{
        "id": t.id,
        "user_id": t.user_id,
        "full_name": t.user.full_name if t.user else None,
        "employee_id": t.employee_id,
        "department": t.department,
        "created_at": t.created_at.isoformat() if t.created_at else None
    } for t in teachers]


@router.get("/groups")
async def get_all_groups(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Barcha guruhlar"""
    await check_admin(current_user, db)
    
    result = await db.execute(
        select(Group)
        .options(selectinload(Group.direction))
        .order_by(Group.name)
    )
    groups = result.scalars().all()
    
    return [{
        "id": g.id,
        "name": g.name,
        "course": g.course,
        "direction_id": g.direction_id,
        "direction_name": g.direction.name if g.direction else None
    } for g in groups]


@router.get("/attendance/report")
async def get_attendance_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    group_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Davomat hisoboti"""
    await check_admin(current_user, db)
    
    query = select(Attendance).options(
        selectinload(Attendance.student).selectinload(Student.user),
        selectinload(Attendance.student).selectinload(Student.group),
        selectinload(Attendance.lesson).selectinload(Lesson.schedule).selectinload(Schedule.subject)
    )
    
    if start_date:
        query = query.join(Lesson).where(Lesson.date >= date.fromisoformat(start_date))
    if end_date:
        query = query.join(Lesson).where(Lesson.date <= date.fromisoformat(end_date))
    if group_id:
        query = query.join(Student).where(Student.group_id == group_id)
    
    result = await db.execute(query.order_by(Attendance.id.desc()).limit(500))
    attendances = result.scalars().all()
    
    return [{
        "id": a.id,
        "student_name": a.student.user.full_name if a.student and a.student.user else None,
        "student_id": a.student.student_id if a.student else None,
        "group_name": a.student.group.name if a.student and a.student.group else None,
        "subject_name": a.lesson.schedule.subject.name if a.lesson and a.lesson.schedule and a.lesson.schedule.subject else None,
        "date": a.lesson.date.isoformat() if a.lesson else None,
        "status": a.status,
        "marked_at": a.marked_at.isoformat() if a.marked_at else None
    } for a in attendances]


@router.get("/attendance/export")
async def export_attendance_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    group_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Davomatni Excel formatda eksport qilish"""
    await check_admin(current_user, db)
    
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    # Ma'lumotlarni olish
    query = select(Attendance).options(
        selectinload(Attendance.student).selectinload(Student.user),
        selectinload(Attendance.student).selectinload(Student.group),
        selectinload(Attendance.lesson).selectinload(Lesson.schedule).selectinload(Schedule.subject)
    )
    
    if start_date:
        start = date.fromisoformat(start_date)
        query = query.join(Lesson).where(Lesson.date >= start)
    if end_date:
        end = date.fromisoformat(end_date)
        query = query.join(Lesson).where(Lesson.date <= end)
    if group_id:
        query = query.join(Student).where(Student.group_id == group_id)
    
    result = await db.execute(query.order_by(Attendance.id.desc()))
    attendances = result.scalars().all()
    
    # Excel yaratish
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Davomat"
    
    # Sarlavhalar
    headers = ["#", "Talaba", "Talaba ID", "Guruh", "Fan", "Sana", "Status", "Vaqt"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Ma'lumotlar
    status_colors = {
        "present": "C6EFCE",
        "late": "FFEB9C", 
        "absent": "FFC7CE"
    }
    
    for row, a in enumerate(attendances, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=a.student.user.full_name if a.student and a.student.user else "-")
        ws.cell(row=row, column=3, value=a.student.student_id if a.student else "-")
        ws.cell(row=row, column=4, value=a.student.group.name if a.student and a.student.group else "-")
        ws.cell(row=row, column=5, value=a.lesson.schedule.subject.name if a.lesson and a.lesson.schedule and a.lesson.schedule.subject else "-")
        ws.cell(row=row, column=6, value=a.lesson.date.isoformat() if a.lesson else "-")
        
        status_cell = ws.cell(row=row, column=7, value=a.status)
        if a.status in status_colors:
            status_cell.fill = PatternFill(start_color=status_colors[a.status], end_color=status_colors[a.status], fill_type="solid")
        
        ws.cell(row=row, column=8, value=a.marked_at.strftime("%H:%M") if a.marked_at else "-")
    
    # Ustun kengliklarini sozlash
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    
    # Faylni saqlash
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"davomat_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/lessons/today")
async def get_today_lessons(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Bugungi darslar"""
    await check_admin(current_user, db)
    
    today = date.today()
    
    result = await db.execute(
        select(Lesson)
        .options(
            selectinload(Lesson.schedule).selectinload(Schedule.subject),
            selectinload(Lesson.schedule).selectinload(Schedule.group),
            selectinload(Lesson.schedule).selectinload(Schedule.teacher).selectinload(Teacher.user)
        )
        .where(Lesson.date == today)
        .order_by(Lesson.id)
    )
    lessons = result.scalars().all()
    
    return [{
        "id": l.id,
        "subject_name": l.schedule.subject.name if l.schedule and l.schedule.subject else None,
        "group_name": l.schedule.group.name if l.schedule and l.schedule.group else None,
        "teacher_name": l.schedule.teacher.user.full_name if l.schedule and l.schedule.teacher and l.schedule.teacher.user else None,
        "status": l.status,
        "opened_at": l.opened_at.isoformat() if l.opened_at else None,
        "closed_at": l.closed_at.isoformat() if l.closed_at else None
    } for l in lessons]


@router.post("/groups/create")
async def create_group(
    name: str,
    direction_id: int,
    course: int = 1,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Yangi guruh qo'shish"""
    await check_admin(current_user, db)
    
    group = Group(name=name, direction_id=direction_id, course=course)
    db.add(group)
    await db.commit()
    
    return {"success": True, "message": "Guruh qo'shildi", "id": group.id}


@router.post("/subjects/create")
async def create_subject(
    name: str,
    short_name: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Yangi fan qo'shish"""
    await check_admin(current_user, db)
    
    subject = Subject(name=name, short_name=short_name)
    db.add(subject)
    await db.commit()
    
    return {"success": True, "message": "Fan qo'shildi", "id": subject.id}


@router.delete("/users/{user_id}")
async def delete_user(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Foydalanuvchini o'chirish"""
    await check_admin(current_user, db)
    
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    
    if not user:
        raise HTTPException(status_code=404, detail="User topilmadi")
    
    await db.delete(user)
    await db.commit()
    
    return {"success": True, "message": "User o'chirildi"}